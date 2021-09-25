from django_bootstrap import bootstrap  # noqa

bootstrap()  # noqa

import sys

from openpyxl import load_workbook

from criteria.models import MajorCuptCode
from appl.models import Faculty

from django.db.utils import IntegrityError

FIELDS = {
    'program_code': 'program_id',
    'program_type': 'program_type_name_th',
    'major_code': 'major_id',
    'title': 'program_name_th',
    'major_title': 'major_name_th',
    'program_type_code': 'program_type_id',
}

FACULTY_FIELD_NAME = 'faculty_name_th'
CAMPUS_FIELD_NAME = 'campus_name_th'

DATA_FIELD_MAP = {}
FACULTY_NAME_IDX = 0
CAMPUS_NAME_IDX = 0

def normalize_faculty_name(faculty_name, campus_name):
    if campus_name == 'กำแพงแสน':
        if faculty_name in ['คณะประมง', 'คณะสิ่งแวดล้อม', 'คณะสัตวแพทยศาสตร์']:
            return f'{faculty_name} {campus_name}'
    return faculty_name
        

def main():
    filename = sys.argv[1]
    counter = 0
    workbook = load_workbook(filename)
    
    sheetname = workbook.sheetnames[0]

    print("Sheet:", sheetname)
    sheet = workbook[sheetname]

    all_rows = list(sheet.rows)
    first_row = all_rows[0]
    rows = all_rows[1:]
    
    print(len(rows),'rows')

    field_map = {
        first_row[idx].value: idx
        for idx in range(len(first_row))
    }

    FACULTY_NAME_IDX = field_map[FACULTY_FIELD_NAME]
    CAMPUS_NAME_IDX = field_map[CAMPUS_FIELD_NAME]
    
    for f in FIELDS:
        DATA_FIELD_MAP[f] = field_map[FIELDS[f]]

    for row in rows:
        try:
            faculty_name = normalize_faculty_name(row[FACULTY_NAME_IDX].value,
                                                  row[CAMPUS_NAME_IDX].value)

            if faculty_name == None:
                continue
            
            values = {
                f: row[DATA_FIELD_MAP[f]].value for f in FIELDS
            }
            print(values)
            print(faculty_name)

            faculty = Faculty.objects.get(title=faculty_name)

            if values['major_code'] == None:
                values['major_code'] = ''
                values['major_title'] = ''
            
            program_code = values['program_code']
            major_code = values['major_code']
            old_codes = MajorCuptCode.objects.filter(program_code=program_code,
                                                     major_code=major_code).all()

            print(faculty_name, values)

            if len(old_codes) != 0:
                major_cupt_code = old_codes[0]
            else:
                major_cupt_code = MajorCuptCode()

            for f in FIELDS:
                setattr(major_cupt_code,f,values[f])
            major_cupt_code.faculty=faculty
                
            major_cupt_code.save()
        except IntegrityError as e:
            s = str(e)
            print("ERROR: %s" % (s))

        counter += 1
    print('Imported', counter, 'major_cupt_code')


if __name__ == "__main__":
    main()
